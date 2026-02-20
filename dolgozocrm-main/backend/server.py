from fastapi import FastAPI, APIRouter, HTTPException, Depends, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
from passlib.context import CryptContext
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Exports directory for Excel files
EXPORTS_DIR = ROOT_DIR / "exports"
EXPORTS_DIR.mkdir(exist_ok=True)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'dolgozocrm-secret-key-2024')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Security
security = HTTPBearer()

app = FastAPI(title="Dolgozó CRM API")
api_router = APIRouter(prefix="/api")

# ==================== MODELS ====================

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: Optional[str] = ""
    role: str = "user"  # "admin" or "user" (toborzó)

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str
    created_at: str

class PasswordChange(BaseModel):
    current_password: str
    new_password: str

class ProfileUpdate(BaseModel):
    name: str

class WorkerTypeCreate(BaseModel):
    name: str

class WorkerTypeResponse(BaseModel):
    id: str
    name: str

class PositionCreate(BaseModel):
    name: str
    worker_type_id: str  # Melyik típushoz tartozik

class PositionResponse(BaseModel):
    id: str
    name: str
    worker_type_id: str
    worker_type_name: Optional[str] = ""

class StatusCreate(BaseModel):
    name: str

class StatusResponse(BaseModel):
    id: str
    name: str

class TagCreate(BaseModel):
    name: str
    color: str = "#6366f1"

class TagResponse(BaseModel):
    id: str
    name: str
    color: str

class WorkerCreate(BaseModel):
    name: str
    phone: str
    worker_type_id: str
    position: Optional[str] = ""  # Szabad szöveg pozíció
    position_experience: Optional[str] = ""  # Pozícióval kapcsolatos tapasztalat
    category: str = "Felvitt dolgozók"
    address: Optional[str] = ""
    email: Optional[str] = ""
    experience: Optional[str] = ""
    notes: Optional[str] = ""

class WorkerUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    worker_type_id: Optional[str] = None
    position: Optional[str] = None
    position_experience: Optional[str] = None
    category: Optional[str] = None
    address: Optional[str] = None
    email: Optional[str] = None
    experience: Optional[str] = None
    notes: Optional[str] = None

class WorkerResponse(BaseModel):
    id: str
    name: str
    phone: str
    worker_type_id: str
    worker_type_name: Optional[str] = ""
    position: Optional[str] = ""
    position_experience: Optional[str] = ""
    category: str
    address: str
    email: str
    experience: str
    notes: str
    tags: List[dict] = []
    project_statuses: List[dict] = []
    owner_id: str
    owner_name: str
    created_at: str

class ProjectCreate(BaseModel):
    name: str
    date: str
    location: Optional[str] = ""
    notes: Optional[str] = ""
    expected_workers: int = 0
    recruiter_ids: List[str] = []  # Hozzárendelt toborzók

class ProjectUpdate(BaseModel):
    name: Optional[str] = None
    date: Optional[str] = None
    location: Optional[str] = None
    notes: Optional[str] = None
    is_closed: Optional[bool] = None
    expected_workers: Optional[int] = None
    recruiter_ids: Optional[List[str]] = None

class ProjectResponse(BaseModel):
    id: str
    name: str
    date: str
    location: str
    notes: str
    is_closed: bool
    worker_count: int
    expected_workers: int
    recruiter_ids: List[str] = []
    recruiters: List[dict] = []
    owner_id: str = ""
    owner_name: str = ""
    created_at: str

class ProjectWorkerAdd(BaseModel):
    worker_id: str
    status_id: Optional[str] = None

class ProjectRecruiterAdd(BaseModel):
    user_id: str

class ProjectWorkerStatusUpdate(BaseModel):
    status_id: str
    notes: Optional[str] = None

class WorkerHistoryEntry(BaseModel):
    project_id: str
    project_name: str
    project_date: str
    status_id: str
    status_name: str
    notes: str
    updated_at: str

# ==================== HELPER FUNCTIONS ====================

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)

def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Felhasználó nem található")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token lejárt")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Érvénytelen token")

async def require_admin(user: dict = Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Csak admin jogosultsággal")
    return user

# ==================== AUTH ENDPOINTS ====================

@api_router.post("/auth/register", response_model=dict)
async def register(data: UserCreate, current_user: dict = Depends(require_admin)):
    """Admin csak hozhat létre új felhasználót"""
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Ez az email már regisztrálva van")
    
    if len(data.password) < 8:
        raise HTTPException(status_code=400, detail="A jelszó minimum 8 karakter legyen")
    
    user_doc = {
        "id": str(uuid.uuid4()),
        "email": data.email,
        "password": hash_password(data.password),
        "name": data.name or data.email.split("@")[0],
        "role": data.role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    return {"message": "Felhasználó létrehozva", "email": data.email}

@api_router.post("/auth/login", response_model=dict)
async def login(data: UserLogin):
    user = await db.users.find_one({"email": data.email}, {"_id": 0})
    if not user or not verify_password(data.password, user["password"]):
        raise HTTPException(status_code=401, detail="Hibás email vagy jelszó")
    
    token = create_token(user["id"], user["email"], user["role"])
    return {
        "token": token,
        "user": {
            "id": user["id"],
            "email": user["email"],
            "name": user["name"],
            "role": user["role"]
        }
    }

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(user: dict = Depends(get_current_user)):
    return UserResponse(
        id=user["id"],
        email=user["email"],
        name=user.get("name", ""),
        role=user["role"],
        created_at=user.get("created_at", "")
    )

@api_router.put("/auth/profile")
async def update_profile(data: ProfileUpdate, user: dict = Depends(get_current_user)):
    await db.users.update_one({"id": user["id"]}, {"$set": {"name": data.name}})
    return {"message": "Profil frissítve"}

@api_router.put("/auth/password")
async def change_password(data: PasswordChange, user: dict = Depends(get_current_user)):
    db_user = await db.users.find_one({"id": user["id"]}, {"_id": 0})
    if not verify_password(data.current_password, db_user["password"]):
        raise HTTPException(status_code=400, detail="Hibás jelenlegi jelszó")
    
    if len(data.new_password) < 8:
        raise HTTPException(status_code=400, detail="Az új jelszó minimum 8 karakter legyen")
    
    await db.users.update_one(
        {"id": user["id"]}, 
        {"$set": {"password": hash_password(data.new_password)}}
    )
    return {"message": "Jelszó megváltoztatva"}

# ==================== WORKER TYPES ====================

@api_router.get("/worker-types", response_model=List[WorkerTypeResponse])
async def get_worker_types(user: dict = Depends(get_current_user)):
    types = await db.worker_types.find({}, {"_id": 0}).to_list(100)
    return [WorkerTypeResponse(**t) for t in types]

@api_router.post("/worker-types", response_model=WorkerTypeResponse)
async def create_worker_type(data: WorkerTypeCreate, user: dict = Depends(require_admin)):
    type_doc = {"id": str(uuid.uuid4()), "name": data.name}
    await db.worker_types.insert_one(type_doc)
    return WorkerTypeResponse(**type_doc)

@api_router.delete("/worker-types/{type_id}")
async def delete_worker_type(type_id: str, user: dict = Depends(require_admin)):
    result = await db.worker_types.delete_one({"id": type_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Típus nem található")
    # Töröljük a típushoz tartozó pozíciókat is
    await db.positions.delete_many({"worker_type_id": type_id})
    return {"message": "Típus törölve"}

# ==================== POSITIONS ====================

@api_router.get("/positions", response_model=List[PositionResponse])
async def get_positions(worker_type_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    """Pozíciók lekérése, opcionálisan típus szerint szűrve"""
    query = {}
    if worker_type_id:
        query["worker_type_id"] = worker_type_id
    
    positions = await db.positions.find(query, {"_id": 0}).to_list(100)
    
    result = []
    for p in positions:
        type_doc = await db.worker_types.find_one({"id": p.get("worker_type_id")}, {"_id": 0})
        result.append(PositionResponse(
            id=p["id"],
            name=p["name"],
            worker_type_id=p["worker_type_id"],
            worker_type_name=type_doc["name"] if type_doc else ""
        ))
    return result

@api_router.post("/positions", response_model=PositionResponse)
async def create_position(data: PositionCreate, user: dict = Depends(require_admin)):
    # Ellenőrizzük, hogy létezik-e a típus
    type_doc = await db.worker_types.find_one({"id": data.worker_type_id}, {"_id": 0})
    if not type_doc:
        raise HTTPException(status_code=404, detail="Típus nem található")
    
    position_doc = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "worker_type_id": data.worker_type_id
    }
    await db.positions.insert_one(position_doc)
    return PositionResponse(**position_doc, worker_type_name=type_doc["name"])

@api_router.delete("/positions/{position_id}")
async def delete_position(position_id: str, user: dict = Depends(require_admin)):
    result = await db.positions.delete_one({"id": position_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pozíció nem található")
    return {"message": "Pozíció törölve"}

# ==================== STATUSES ====================

@api_router.get("/statuses", response_model=List[StatusResponse])
async def get_statuses(user: dict = Depends(get_current_user)):
    statuses = await db.statuses.find({}, {"_id": 0}).to_list(100)
    return [StatusResponse(**s) for s in statuses]

@api_router.post("/statuses", response_model=StatusResponse)
async def create_status(data: StatusCreate, user: dict = Depends(require_admin)):
    status_doc = {"id": str(uuid.uuid4()), "name": data.name}
    await db.statuses.insert_one(status_doc)
    return StatusResponse(**status_doc)

@api_router.delete("/statuses/{status_id}")
async def delete_status(status_id: str, user: dict = Depends(require_admin)):
    result = await db.statuses.delete_one({"id": status_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Státusz nem található")
    return {"message": "Státusz törölve"}

# ==================== TAGS ====================

@api_router.get("/tags", response_model=List[TagResponse])
async def get_tags(user: dict = Depends(get_current_user)):
    tags = await db.tags.find({}, {"_id": 0}).to_list(100)
    return [TagResponse(**t) for t in tags]

@api_router.post("/tags", response_model=TagResponse)
async def create_tag(data: TagCreate, user: dict = Depends(require_admin)):
    tag_doc = {"id": str(uuid.uuid4()), "name": data.name, "color": data.color}
    await db.tags.insert_one(tag_doc)
    return TagResponse(**tag_doc)

@api_router.delete("/tags/{tag_id}")
async def delete_tag(tag_id: str, user: dict = Depends(require_admin)):
    result = await db.tags.delete_one({"id": tag_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Jellemző nem található")
    return {"message": "Jellemző törölve"}

# ==================== USERS (Admin) ====================

@api_router.get("/users", response_model=List[UserResponse])
async def get_users(user: dict = Depends(require_admin)):
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(100)
    return [UserResponse(
        id=u["id"],
        email=u["email"],
        name=u.get("name", ""),
        role=u["role"],
        created_at=u.get("created_at", "")
    ) for u in users]

@api_router.get("/users/stats")
async def get_user_stats(user: dict = Depends(require_admin)):
    """Toborzónként hány dolgozót vitt fel"""
    pipeline = [
        {"$group": {"_id": "$owner_id", "count": {"$sum": 1}}},
    ]
    stats = await db.workers.aggregate(pipeline).to_list(100)
    
    result = []
    for s in stats:
        owner = await db.users.find_one({"id": s["_id"]}, {"_id": 0, "password": 0})
        if owner:
            result.append({
                "user_id": s["_id"],
                "user_name": owner.get("name", owner["email"]),
                "user_email": owner["email"],
                "worker_count": s["count"]
            })
    return result

# ==================== WORKERS ====================

@api_router.get("/workers", response_model=List[WorkerResponse])
async def get_workers(
    search: Optional[str] = None,
    category: Optional[str] = None,
    worker_type_id: Optional[str] = None,
    tag_id: Optional[str] = None,
    owner_id: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    
    # Toborzó csak saját dolgozóit látja
    if user["role"] != "admin":
        query["owner_id"] = user["id"]
    elif owner_id:
        query["owner_id"] = owner_id
    
    if category:
        query["category"] = category
    if worker_type_id:
        query["worker_type_id"] = worker_type_id
    if tag_id:
        query["tag_ids"] = tag_id
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"address": {"$regex": search, "$options": "i"}},
            {"experience": {"$regex": search, "$options": "i"}},
            {"position": {"$regex": search, "$options": "i"}}
        ]
    
    workers = await db.workers.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Enrich with type names, tags, project statuses
    result = []
    for w in workers:
        # Get type name
        type_doc = await db.worker_types.find_one({"id": w.get("worker_type_id")}, {"_id": 0})
        w["worker_type_name"] = type_doc["name"] if type_doc else ""
        
        # Position is now free text
        w["position"] = w.get("position", "")
        w["position_experience"] = w.get("position_experience", "")
        
        # Get tags
        tag_ids = w.get("tag_ids", [])
        tags = []
        for tid in tag_ids:
            tag = await db.tags.find_one({"id": tid}, {"_id": 0})
            if tag:
                tags.append(tag)
        w["tags"] = tags
        
        # Get project statuses
        project_workers = await db.project_workers.find(
            {"worker_id": w["id"]}, {"_id": 0}
        ).sort("updated_at", -1).to_list(100)
        
        project_statuses = []
        for pw in project_workers:
            project = await db.projects.find_one({"id": pw["project_id"]}, {"_id": 0})
            status = await db.statuses.find_one({"id": pw.get("status_id")}, {"_id": 0})
            if project:
                project_statuses.append({
                    "project_id": project["id"],
                    "project_name": project["name"],
                    "project_date": project.get("date", ""),
                    "status_id": pw.get("status_id", ""),
                    "status_name": status["name"] if status else "Hozzárendelve",
                    "notes": pw.get("notes", ""),
                    "updated_at": pw.get("updated_at", "")
                })
        w["project_statuses"] = project_statuses
        
        # Get owner name
        owner = await db.users.find_one({"id": w.get("owner_id")}, {"_id": 0, "password": 0})
        w["owner_name"] = owner.get("name", owner["email"]) if owner else ""
        
        result.append(WorkerResponse(**w))
    
    return result

@api_router.get("/workers/{worker_id}", response_model=WorkerResponse)
async def get_worker(worker_id: str, user: dict = Depends(get_current_user)):
    query = {"id": worker_id}
    if user["role"] != "admin":
        query["owner_id"] = user["id"]
    
    w = await db.workers.find_one(query, {"_id": 0})
    if not w:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    # Enrich
    type_doc = await db.worker_types.find_one({"id": w.get("worker_type_id")}, {"_id": 0})
    w["worker_type_name"] = type_doc["name"] if type_doc else ""
    
    # Position is now free text
    w["position"] = w.get("position", "")
    w["position_experience"] = w.get("position_experience", "")
    
    tag_ids = w.get("tag_ids", [])
    tags = []
    for tid in tag_ids:
        tag = await db.tags.find_one({"id": tid}, {"_id": 0})
        if tag:
            tags.append(tag)
    w["tags"] = tags
    
    project_workers = await db.project_workers.find(
        {"worker_id": w["id"]}, {"_id": 0}
    ).sort("updated_at", -1).to_list(100)
    
    project_statuses = []
    for pw in project_workers:
        project = await db.projects.find_one({"id": pw["project_id"]}, {"_id": 0})
        status = await db.statuses.find_one({"id": pw.get("status_id")}, {"_id": 0})
        if project:
            project_statuses.append({
                "project_id": project["id"],
                "project_name": project["name"],
                "project_date": project.get("date", ""),
                "status_id": pw.get("status_id", ""),
                "status_name": status["name"] if status else "Hozzárendelve",
                "notes": pw.get("notes", ""),
                "updated_at": pw.get("updated_at", "")
            })
    w["project_statuses"] = project_statuses
    
    owner = await db.users.find_one({"id": w.get("owner_id")}, {"_id": 0, "password": 0})
    w["owner_name"] = owner.get("name", owner["email"]) if owner else ""
    
    return WorkerResponse(**w)

@api_router.post("/workers", response_model=WorkerResponse)
async def create_worker(data: WorkerCreate, user: dict = Depends(get_current_user)):
    if len(data.name) < 2:
        raise HTTPException(status_code=400, detail="A név minimum 2 karakter legyen")
    
    worker_doc = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "phone": data.phone,
        "worker_type_id": data.worker_type_id,
        "position": data.position or "",
        "position_experience": data.position_experience or "",
        "category": data.category,
        "address": data.address or "",
        "email": data.email or "",
        "experience": data.experience or "",
        "notes": data.notes or "",
        "tag_ids": [],
        "owner_id": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.workers.insert_one(worker_doc)
    
    worker_doc["worker_type_name"] = ""
    worker_doc["tags"] = []
    worker_doc["project_statuses"] = []
    worker_doc["owner_name"] = user.get("name", user["email"])
    
    return WorkerResponse(**worker_doc)

@api_router.put("/workers/{worker_id}", response_model=WorkerResponse)
async def update_worker(worker_id: str, data: WorkerUpdate, user: dict = Depends(get_current_user)):
    query = {"id": worker_id}
    if user["role"] != "admin":
        query["owner_id"] = user["id"]
    
    worker = await db.workers.find_one(query, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.workers.update_one({"id": worker_id}, {"$set": update_data})
    
    return await get_worker(worker_id, user)

@api_router.delete("/workers/{worker_id}")
async def delete_worker(worker_id: str, user: dict = Depends(require_admin)):
    """Csak admin törölhet"""
    result = await db.workers.delete_one({"id": worker_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    # Töröljük a projekt kapcsolatokat is
    await db.project_workers.delete_many({"worker_id": worker_id})
    
    return {"message": "Dolgozó törölve"}

@api_router.post("/workers/{worker_id}/tags/{tag_id}")
async def add_tag_to_worker(worker_id: str, tag_id: str, user: dict = Depends(get_current_user)):
    query = {"id": worker_id}
    if user["role"] != "admin":
        query["owner_id"] = user["id"]
    
    worker = await db.workers.find_one(query)
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$addToSet": {"tag_ids": tag_id}}
    )
    return {"message": "Jellemző hozzáadva"}

@api_router.delete("/workers/{worker_id}/tags/{tag_id}")
async def remove_tag_from_worker(worker_id: str, tag_id: str, user: dict = Depends(get_current_user)):
    query = {"id": worker_id}
    if user["role"] != "admin":
        query["owner_id"] = user["id"]
    
    worker = await db.workers.find_one(query)
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$pull": {"tag_ids": tag_id}}
    )
    return {"message": "Jellemző eltávolítva"}

# ==================== PROJECTS ====================

@api_router.get("/projects", response_model=List[ProjectResponse])
async def get_projects(user: dict = Depends(get_current_user)):
    """Toborzó csak azokat a projekteket látja, ahol ő hozta létre VAGY hozzá van rendelve"""
    projects = await db.projects.find({}, {"_id": 0}).sort("date", -1).to_list(1000)
    
    result = []
    for p in projects:
        count = await db.project_workers.count_documents({"project_id": p["id"]})
        
        recruiter_ids = p.get("recruiter_ids", [])
        owner_id = p.get("owner_id", "")
        
        # Ha toborzó, csak azokat mutassa ahol ő hozta létre VAGY hozzá van rendelve
        if user["role"] != "admin":
            if owner_id != user["id"] and user["id"] not in recruiter_ids:
                continue
        
        # Get recruiter names
        recruiters = []
        for rid in recruiter_ids:
            r = await db.users.find_one({"id": rid}, {"_id": 0, "password": 0})
            if r:
                recruiters.append({"id": r["id"], "name": r.get("name", r["email"]), "email": r["email"]})
        
        # Get owner name
        owner_name = ""
        if owner_id:
            owner = await db.users.find_one({"id": owner_id}, {"_id": 0, "password": 0})
            if owner:
                owner_name = owner.get("name", owner["email"])
        
        result.append(ProjectResponse(
            id=p["id"],
            name=p["name"],
            date=p["date"],
            location=p.get("location", ""),
            notes=p.get("notes", ""),
            is_closed=p.get("is_closed", False),
            worker_count=count,
            expected_workers=p.get("expected_workers", 0),
            recruiter_ids=recruiter_ids,
            recruiters=recruiters,
            owner_id=owner_id,
            owner_name=owner_name,
            created_at=p.get("created_at", "")
        ))
    
    return result

@api_router.get("/projects/{project_id}")
async def get_project(project_id: str, user: dict = Depends(get_current_user)):
    p = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    # Ellenőrizzük jogosultságot
    recruiter_ids = p.get("recruiter_ids", [])
    owner_id = p.get("owner_id", "")
    if user["role"] != "admin" and owner_id != user["id"] and user["id"] not in recruiter_ids:
        raise HTTPException(status_code=403, detail="Nincs hozzáférésed ehhez a projekthez")
    
    # Get workers
    pw_list = await db.project_workers.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    
    workers = []
    for pw in pw_list:
        w = await db.workers.find_one({"id": pw["worker_id"]}, {"_id": 0})
        if w:
            # Toborzó csak saját dolgozóit látja a projektben
            if user["role"] != "admin" and w.get("owner_id") != user["id"]:
                continue
            status = await db.statuses.find_one({"id": pw.get("status_id")}, {"_id": 0})
            type_doc = await db.worker_types.find_one({"id": w.get("worker_type_id")}, {"_id": 0})
            owner = await db.users.find_one({"id": w.get("owner_id")}, {"_id": 0, "password": 0})
            workers.append({
                "id": w["id"],
                "name": w["name"],
                "phone": w["phone"],
                "category": w["category"],
                "worker_type_name": type_doc["name"] if type_doc else "",
                "status_id": pw.get("status_id", ""),
                "status_name": status["name"] if status else "Hozzárendelve",
                "notes": pw.get("notes", ""),
                "added_by": owner.get("name", owner["email"]) if owner else "",
                "added_at": pw.get("created_at", "")
            })
    
    total_count = await db.project_workers.count_documents({"project_id": project_id})
    
    # Get recruiter names
    recruiters = []
    for rid in recruiter_ids:
        r = await db.users.find_one({"id": rid}, {"_id": 0, "password": 0})
        if r:
            recruiters.append({"id": r["id"], "name": r.get("name", r["email"]), "email": r["email"]})
    
    # Get owner name
    owner_name = ""
    if owner_id:
        owner = await db.users.find_one({"id": owner_id}, {"_id": 0, "password": 0})
        if owner:
            owner_name = owner.get("name", owner["email"])
    
    return {
        "id": p["id"],
        "name": p["name"],
        "date": p["date"],
        "location": p.get("location", ""),
        "notes": p.get("notes", ""),
        "is_closed": p.get("is_closed", False),
        "worker_count": total_count,
        "expected_workers": p.get("expected_workers", 0),
        "recruiter_ids": recruiter_ids,
        "recruiters": recruiters,
        "owner_id": owner_id,
        "owner_name": owner_name,
        "workers": workers,
        "created_at": p.get("created_at", "")
    }

@api_router.post("/projects", response_model=ProjectResponse)
async def create_project(data: ProjectCreate, user: dict = Depends(require_admin)):
    """Csak admin hozhat létre projektet"""
    project_doc = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "date": data.date,
        "location": data.location or "",
        "notes": data.notes or "",
        "expected_workers": data.expected_workers,
        "recruiter_ids": data.recruiter_ids,  # Hozzárendelt toborzók
        "is_closed": False,
        "owner_id": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.projects.insert_one(project_doc)
    
    owner_name = user.get("name", user["email"])
    return ProjectResponse(**project_doc, worker_count=0, recruiters=[], owner_name=owner_name)

@api_router.put("/projects/{project_id}", response_model=ProjectResponse)
async def update_project(project_id: str, data: ProjectUpdate, user: dict = Depends(get_current_user)):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.projects.update_one({"id": project_id}, {"$set": update_data})
    
    updated = await db.projects.find_one({"id": project_id}, {"_id": 0})
    count = await db.project_workers.count_documents({"project_id": project_id})
    
    # Get recruiters
    recruiters = []
    for rid in updated.get("recruiter_ids", []):
        r = await db.users.find_one({"id": rid}, {"_id": 0, "password": 0})
        if r:
            recruiters.append({"id": r["id"], "name": r.get("name", r["email"]), "email": r["email"]})
    
    # Get owner name
    owner_name = ""
    owner_id = updated.get("owner_id", "")
    if owner_id:
        owner = await db.users.find_one({"id": owner_id}, {"_id": 0, "password": 0})
        if owner:
            owner_name = owner.get("name", owner["email"])
    
    return ProjectResponse(**updated, worker_count=count, recruiters=recruiters, owner_name=owner_name)

@api_router.post("/projects/{project_id}/recruiters")
async def add_recruiter_to_project(project_id: str, data: ProjectRecruiterAdd, user: dict = Depends(require_admin)):
    """Admin hozzárendel egy toborzót a projekthez"""
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    target_user = await db.users.find_one({"id": data.user_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="Felhasználó nem található")
    
    await db.projects.update_one(
        {"id": project_id},
        {"$addToSet": {"recruiter_ids": data.user_id}}
    )
    return {"message": "Toborzó hozzárendelve a projekthez"}

@api_router.delete("/projects/{project_id}/recruiters/{user_id}")
async def remove_recruiter_from_project(project_id: str, user_id: str, user: dict = Depends(require_admin)):
    """Admin eltávolít egy toborzót a projektből"""
    result = await db.projects.update_one(
        {"id": project_id},
        {"$pull": {"recruiter_ids": user_id}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    return {"message": "Toborzó eltávolítva a projektről"}

@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str, user: dict = Depends(require_admin)):
    """Csak admin törölhet projektet"""
    result = await db.projects.delete_one({"id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    await db.project_workers.delete_many({"project_id": project_id})
    return {"message": "Projekt törölve"}

@api_router.post("/projects/{project_id}/workers")
async def add_worker_to_project(project_id: str, data: ProjectWorkerAdd, user: dict = Depends(get_current_user)):
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    worker = await db.workers.find_one({"id": data.worker_id})
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    existing = await db.project_workers.find_one({
        "project_id": project_id,
        "worker_id": data.worker_id
    })
    if existing:
        raise HTTPException(status_code=400, detail="Dolgozó már hozzá van rendelve")
    
    pw_doc = {
        "id": str(uuid.uuid4()),
        "project_id": project_id,
        "worker_id": data.worker_id,
        "status_id": data.status_id or "",
        "added_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.project_workers.insert_one(pw_doc)
    return {"message": "Dolgozó hozzáadva a projekthez"}

@api_router.delete("/projects/{project_id}/workers/{worker_id}")
async def remove_worker_from_project(project_id: str, worker_id: str, user: dict = Depends(get_current_user)):
    result = await db.project_workers.delete_one({
        "project_id": project_id,
        "worker_id": worker_id
    })
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kapcsolat nem található")
    return {"message": "Dolgozó eltávolítva a projektről"}

@api_router.put("/projects/{project_id}/workers/{worker_id}/status")
async def update_worker_status_in_project(
    project_id: str, 
    worker_id: str, 
    data: ProjectWorkerStatusUpdate,
    user: dict = Depends(get_current_user)
):
    update_fields = {
        "status_id": data.status_id,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    if data.notes is not None:
        update_fields["notes"] = data.notes
    
    result = await db.project_workers.update_one(
        {"project_id": project_id, "worker_id": worker_id},
        {"$set": update_fields}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Kapcsolat nem található")
    return {"message": "Státusz frissítve"}

# ==================== EXCEL EXPORT ====================

async def generate_excel_for_user(user_id: str, user_name: str):
    """Generate Excel file for a specific recruiter with workers grouped by category"""
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    categories = ["Felvitt dolgozók", "Hideg jelentkező", "Űrlapon jelentkezett", 
                  "Állásra jelentkezett", "Ingázó", "Szállásos"]
    
    # Remove default sheet
    wb.remove(wb.active)
    
    for cat in categories:
        workers = await db.workers.find(
            {"owner_id": user_id, "category": cat}, {"_id": 0}
        ).sort("name", 1).to_list(1000)
        
        if not workers:
            continue
            
        # Create sheet for category
        ws = wb.create_sheet(title=cat[:31])  # Excel max 31 chars
        
        # Headers
        headers = ["Név", "Telefon", "Email", "Lakcím", "Típus", "Tapasztalat", "Felvéve"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row, worker in enumerate(workers, 2):
            type_doc = await db.worker_types.find_one({"id": worker.get("worker_type_id")}, {"_id": 0})
            type_name = type_doc["name"] if type_doc else ""
            
            ws.cell(row=row, column=1, value=worker["name"]).border = border
            ws.cell(row=row, column=2, value=worker["phone"]).border = border
            ws.cell(row=row, column=3, value=worker.get("email", "")).border = border
            ws.cell(row=row, column=4, value=worker.get("address", "")).border = border
            ws.cell(row=row, column=5, value=type_name).border = border
            ws.cell(row=row, column=6, value=worker.get("experience", "")).border = border
            ws.cell(row=row, column=7, value=worker.get("created_at", "")[:10]).border = border
        
        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    # If no sheets were created, add summary
    if not wb.sheetnames:
        ws = wb.create_sheet(title="Összefoglaló")
        ws.cell(row=1, column=1, value="Nincs dolgozó ebben a kategóriában")
    
    # Save file
    safe_name = "".join(c for c in user_name if c.isalnum() or c in " -_").strip() or "export"
    filename = f"{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = EXPORTS_DIR / filename
    wb.save(filepath)
    
    return filepath, filename

@api_router.get("/export/workers")
async def export_workers_excel(user: dict = Depends(get_current_user)):
    """Export current user's workers to Excel"""
    user_name = user.get("name") or user["email"].split("@")[0]
    filepath, filename = await generate_excel_for_user(user["id"], user_name)
    
    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@api_router.get("/export/workers/{user_id}")
async def export_user_workers_excel(user_id: str, admin: dict = Depends(require_admin)):
    """Admin can export any user's workers to Excel"""
    target_user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="Felhasználó nem található")
    
    user_name = target_user.get("name") or target_user["email"].split("@")[0]
    filepath, filename = await generate_excel_for_user(user_id, user_name)
    
    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@api_router.get("/export/all")
async def export_all_workers_excel(admin: dict = Depends(require_admin)):
    """Admin exports all workers grouped by recruiter and category"""
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    recruiter_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get all users
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(100)
    
    # Remove default sheet
    wb.remove(wb.active)
    
    for u in users:
        workers = await db.workers.find({"owner_id": u["id"]}, {"_id": 0}).sort("category", 1).to_list(1000)
        
        if not workers:
            continue
        
        user_name = u.get("name") or u["email"].split("@")[0]
        sheet_name = user_name[:31]  # Excel max 31 chars
        
        # Handle duplicate sheet names
        if sheet_name in wb.sheetnames:
            sheet_name = f"{sheet_name[:28]}_{len(wb.sheetnames)}"
        
        ws = wb.create_sheet(title=sheet_name)
        
        # Headers
        headers = ["Név", "Telefon", "Email", "Kategória", "Típus", "Lakcím", "Felvéve"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Data
        for row, worker in enumerate(workers, 2):
            type_doc = await db.worker_types.find_one({"id": worker.get("worker_type_id")}, {"_id": 0})
            type_name = type_doc["name"] if type_doc else ""
            
            ws.cell(row=row, column=1, value=worker["name"]).border = border
            ws.cell(row=row, column=2, value=worker["phone"]).border = border
            ws.cell(row=row, column=3, value=worker.get("email", "")).border = border
            ws.cell(row=row, column=4, value=worker["category"]).border = border
            ws.cell(row=row, column=5, value=type_name).border = border
            ws.cell(row=row, column=6, value=worker.get("address", "")).border = border
            ws.cell(row=row, column=7, value=worker.get("created_at", "")[:10]).border = border
        
        # Auto-width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    if not wb.sheetnames:
        ws = wb.create_sheet(title="Összefoglaló")
        ws.cell(row=1, column=1, value="Nincs dolgozó a rendszerben")
    
    filename = f"osszes_dolgozo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = EXPORTS_DIR / filename
    wb.save(filepath)
    
    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ==================== SEED DATA ====================

@api_router.post("/seed")
async def seed_data():
    """Initialize default data"""
    # Check if already seeded
    admin = await db.users.find_one({"email": "admin@dolgozocrm.hu"})
    if admin:
        return {"message": "Adatok már léteznek"}
    
    # Create admin user
    admin_doc = {
        "id": str(uuid.uuid4()),
        "email": "admin@dolgozocrm.hu",
        "password": hash_password("admin123"),
        "name": "Admin",
        "role": "admin",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(admin_doc)
    
    # Create test recruiter
    recruiter_doc = {
        "id": str(uuid.uuid4()),
        "email": "toborzo@dolgozocrm.hu",
        "password": hash_password("toborzo123"),
        "name": "Teszt Toborzó",
        "role": "user",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(recruiter_doc)
    
    # Worker types with positions
    type_positions = {
        "Betanított munkás": ["Csomagoló", "Komissiózó", "Összeszerelő", "Gyártósori munkás"],
        "Szakmunkás": ["Hegesztő", "Villanyszerelő", "Lakatos", "Esztergályos", "CNC gépkezelő", "Szerszámkészítő"],
        "Targoncás": ["Homlok targoncás", "Oldal targoncás", "Reach truck kezelő", "Magasraktári targoncás"],
        "Gépkezelő": ["Présgép kezelő", "Fröccsöntő gép kezelő", "Hajlítógép kezelő", "Varrógép kezelő"],
        "Raktáros": ["Áruátvevő", "Kiadó", "Leltáros", "Raktári adminisztrátor"],
        "Segédmunkás": ["Takarító", "Anyagmozgató", "Betanított segéd", "Kézi rakodó"]
    }
    
    for type_name, positions in type_positions.items():
        type_id = str(uuid.uuid4())
        await db.worker_types.insert_one({"id": type_id, "name": type_name})
        for pos in positions:
            await db.positions.insert_one({
                "id": str(uuid.uuid4()),
                "name": pos,
                "worker_type_id": type_id
            })
    
    # Statuses
    statuses = ["Jelentkezett", "Megerősítve", "Dolgozik", "Megfelelt", "Nem felelt meg", "Lemondta", "Nem jelent meg"]
    for s in statuses:
        await db.statuses.insert_one({"id": str(uuid.uuid4()), "name": s})
    
    # Tags
    tags = [
        {"name": "Megbízható", "color": "#22c55e"},
        {"name": "Tapasztalt", "color": "#3b82f6"},
        {"name": "Ajánlott", "color": "#f97316"},
        {"name": "Saját autó", "color": "#8b5cf6"},
        {"name": "Éjszakás", "color": "#6366f1"}
    ]
    for t in tags:
        await db.tags.insert_one({"id": str(uuid.uuid4()), **t})
    
    return {"message": "Seed adatok létrehozva", "admin_email": "admin@dolgozocrm.hu", "admin_password": "admin123"}

# ==================== HEALTH ====================

@api_router.get("/")
async def root():
    return {"message": "Dolgozó CRM API", "status": "running"}

@api_router.get("/health")
async def health():
    return {"status": "healthy"}

# ==================== APP CONFIG ====================

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
